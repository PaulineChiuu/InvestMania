import matplotlib
matplotlib.use('Agg')

from flask import Flask, render_template, request, jsonify, redirect, url_for
import yfinance as yf
import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler
from tensorflow.keras.models import Sequential, load_model
from tensorflow.keras.layers import LSTM, Dense, Dropout
from tensorflow.keras.callbacks import EarlyStopping  # 添加這行
import tensorflow as tf
from collections import OrderedDict
import matplotlib.pyplot as plt
import seaborn as sns
import io
import base64
from datetime import datetime, timedelta
import logging
import os
import json
from datetime import datetime
import logging
from stock_recorder import StockRecorder
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 設置日誌
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 添加模型保存路徑
MODEL_PATH = 'models'
if not os.path.exists(MODEL_PATH):
    os.makedirs(MODEL_PATH)

def fetch_stock_data(symbol, period='2y'):
    """從Yahoo Finance獲取股票數據"""
    try:
        logger.info(f"開始獲取股票 {symbol} 的數據")
        stock = yf.Ticker(symbol)
        data = stock.history(period=period)
        
        if data.empty:
            logger.error(f"獲取到的股票 {symbol} 數據為空")
            return None
            
        logger.info(f"成功獲取到 {len(data)} 筆數據")
        data = data.reset_index()
        data['Date'] = pd.to_datetime(data['Date'])
        
        # 添加技術指標
        data['MA5'] = data['Close'].rolling(window=5).mean()
        data['MA20'] = data['Close'].rolling(window=20).mean()
        data['RSI'] = calculate_rsi(data['Close'], 14)
        data['MACD'], data['Signal'] = calculate_macd(data['Close'])
        
        return data
        
    except Exception as e:
        logger.error(f"獲取股票數據時發生錯誤: {str(e)}")
        return None
    
def plot_stock_trend(data):
    """繪製美化後的股票趨勢圖"""
    try:
        logger.info("開始繪製趨勢圖")
        
        # 設置中文字體
        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'Microsoft YaHei', 'SimHei', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False  # 解決負號顯示問題
        
        # 確保 matplotlib 使用 Agg 後端
        matplotlib.use('Agg')
        
        # 創建圖形和軸對象
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # 繪製主要價格線
        ax.plot(data['Date'], data['Close'], 
               label='收盤價', 
               color='#1f77b4', 
               linewidth=2,
               alpha=0.8)
        
        # 添加MA5和MA20移動平均線
        ax.plot(data['Date'], data['MA5'],
               label='5日均線',
               color='#2ca02c',
               linewidth=1,
               alpha=0.8)
        
        ax.plot(data['Date'], data['MA20'],
               label='20日均線',
               color='#ff7f0e',
               linewidth=1,
               alpha=0.8)
        
        # 添加漸層填充
        ax.fill_between(data['Date'], 
                       data['Close'].min(), 
                       data['Close'],
                       alpha=0.1,
                       color='#1f77b4')
        
        # 設置標題和標籤
        plt.title('股票價格趨勢圖', fontsize=16, pad=20)
        plt.xlabel('日期', fontsize=12)
        plt.ylabel('價格', fontsize=12)
        
        # 添加圖例
        plt.legend(loc='upper left', frameon=True)
        
        # 美化刻度
        ax.tick_params(axis='both', labelsize=10)
        plt.xticks(rotation=45)
        
        # 添加網格
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # 調整佈局
        plt.tight_layout()
        
        # 保存圖形前清除先前的圖形
        plt.figure(fig.number)
        
        # 保存圖形
        img = io.BytesIO()
        plt.savefig(img, format='png', bbox_inches='tight', dpi=300)
        img.seek(0)
        img_b64 = base64.b64encode(img.getvalue()).decode('utf-8')
        
        # 清理
        plt.close('all')
        
        logger.info("趨勢圖繪製完成")
        return img_b64
        
    except Exception as e:
        logger.error(f"繪製趨勢圖時發生錯誤: {str(e)}")
        raise

def calculate_rsi(prices, period=14):
    """計算RSI指標"""
    delta = prices.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))

def calculate_macd(prices, slow=26, fast=12, signal=9):
    """計算MACD指標"""
    exp1 = prices.ewm(span=fast, adjust=False).mean()
    exp2 = prices.ewm(span=slow, adjust=False).mean()
    macd = exp1 - exp2
    signal_line = macd.ewm(span=signal, adjust=False).mean()
    return macd, signal_line
def get_state(change):
    #"""根據變化百分比判斷狀態"""
        if change > 0.06:
            return "++"
        elif change > 0.005:
            return "+"
        elif change > -0.005:
            return "0"
        elif change > -0.06:
            return "-"
        else:
            return "--"
class StockPredictor:
    def __init__(self, data, symbol):
        self.data = data
        self.symbol = symbol
        self.scaler = MinMaxScaler(feature_range=(0, 1))
        self.feature_scaler = MinMaxScaler(feature_range=(0, 1))
        self.lstm_model = None
        self.sequence_length = 20
        self.markov_states = None
        self.transition_matrix = None
        self.training_history = None
        
        # 定義固定的特徵列表
        self.features = [
            'Returns', 'Volatility', 
            'Price_MA5_Diff', 'Price_MA20_Diff',
            'RSI', 'BB_width', 
            'MACD_Hist', 'Volume_Ratio',
            'Markov_State'
        ]
        
        # 定義狀態順序
        self.ordered_states = ['大漲', '小漲', '平盤', '小跌', '大跌']
        
        # 定義狀態閾值（使用OrderedDict保持順序）
        self.state_thresholds = OrderedDict([
            ('大漲', 0.06),      # 6%以上
            ('小漲', 0.005),     # 0.5%到6%
            ('平盤', -0.005),    # -0.5%到0.5%
            ('小跌', -0.06),     # -6%到-0.5%
            ('大跌', float('-inf'))  # -6%以下
        ])
        
        # 狀態符號映射
        self.state_symbols = OrderedDict([
            ('大漲', '++'),
            ('小漲', '+'),
            ('平盤', '~'),
            ('小跌', '-'),
            ('大跌', '--')
        ])

    def prepare_features(self):
        """準備預測所需的特徵"""
        try:
            df = self.data.copy()
            
            # 基本價格特徵
            df['Returns'] = df['Close'].pct_change()
            df['Volatility'] = df['Returns'].rolling(window=20).std()
            
            # 移動平均
            df['MA5'] = df['Close'].rolling(window=5).mean()
            df['MA20'] = df['Close'].rolling(window=20).mean()
            df['MA60'] = df['Close'].rolling(window=60).mean()
            
            # 價格與均線差距
            df['Price_MA5_Diff'] = (df['Close'] - df['MA5']) / df['MA5']
            df['Price_MA20_Diff'] = (df['Close'] - df['MA20']) / df['MA20']
            
            # RSI
            delta = df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            df['RSI'] = 100 - (100 / (1 + rs))
            
            # 布林通道
            df['BB_middle'] = df['Close'].rolling(window=20).mean()
            df['BB_upper'] = df['BB_middle'] + 2 * df['Close'].rolling(window=20).std()
            df['BB_lower'] = df['BB_middle'] - 2 * df['Close'].rolling(window=20).std()
            df['BB_width'] = (df['BB_upper'] - df['BB_lower']) / df['BB_middle']
            
            # MACD
            exp1 = df['Close'].ewm(span=12, adjust=False).mean()
            exp2 = df['Close'].ewm(span=26, adjust=False).mean()
            df['MACD'] = exp1 - exp2
            df['Signal_Line'] = df['MACD'].ewm(span=9, adjust=False).mean()
            df['MACD_Hist'] = df['MACD'] - df['Signal_Line']
            
            # 成交量指標
            df['Volume_MA5'] = df['Volume'].rolling(window=5).mean()
            df['Volume_MA20'] = df['Volume'].rolling(window=20).mean()
            df['Volume_Ratio'] = df['Volume'] / df['Volume_MA20']
            
            # 添加馬可夫狀態
            def get_state(returns):
                if returns > self.state_thresholds['大漲']:
                    return 4  # 大漲
                elif returns > self.state_thresholds['小漲']:
                    return 3  # 小漲
                elif returns > self.state_thresholds['平盤']:
                    return 2  # 平盤
                elif returns > self.state_thresholds['小跌']:
                    return 1  # 小跌
                else:
                    return 0  # 大跌
            
            df['Markov_State'] = df['Returns'].apply(get_state)
            
            # 確保回傳的數據框包含所有需要的特徵
            required_features = [
                'Returns', 'Volatility', 
                'Price_MA5_Diff', 'Price_MA20_Diff',
                'RSI', 'BB_width', 
                'MACD_Hist', 'Volume_Ratio',
                'Markov_State'
            ]
            
            # 檢查是否所有需要的特徵都存在
            for feature in required_features:
                if feature not in df.columns:
                    raise ValueError(f"缺少特徵: {feature}")
                    
            return df.dropna()
            
        except Exception as e:
            logger.error(f"準備特徵時發生錯誤: {str(e)}")
            raise

    def prepare_lstm_data(self):
        """准备LSTM训练数据"""
        try:
            # 准备特征数据
            df = self.prepare_features()
            
            # 确保特征存在
            features = [
                'Returns', 'Volatility', 
                'Price_MA5_Diff', 'Price_MA20_Diff',
                'RSI', 'BB_width', 
                'MACD_Hist', 'Volume_Ratio',
                'Markov_State'
            ]
            missing_features = set(features) - set(df.columns)
            if missing_features:
                raise ValueError(f"缺少以下特征: {missing_features}")
            
            # 提取并缩放特征数据
            feature_data = df[features].values
            if not hasattr(self, 'feature_scaler_fitted'):
                scaled_features = self.feature_scaler.fit_transform(feature_data)
                self.feature_scaler_fitted = True
            else:
                scaled_features = self.feature_scaler.transform(feature_data)
            
            # 调用prepare_markov_chain来生成转移矩阵
            if not hasattr(self, 'transition_matrix') or self.transition_matrix is None:
                self.prepare_markov_chain()
                
            if self.transition_matrix is None:
                raise ValueError("转移矩阵生成失败")
            
            # 计算马尔可夫预测值
            markov_predictions = []
            for i in range(len(df) - 1):
                current_state = int(df['Markov_State'].iloc[i])
                predicted_state_probs = self.transition_matrix[current_state]
                predicted_state = np.argmax(predicted_state_probs)  # 取机率最高的状态
                markov_predictions.append(predicted_state)
            markov_predictions = np.array(markov_predictions)
            
            # 准备 X 和 y
            X, y = [], []
            for i in range(self.sequence_length, len(scaled_features) - 1):
                X.append(scaled_features[i - self.sequence_length:i])
                
                # 目标值为实际值与马尔可夫预测的差异
                actual_return = df['Returns'].iloc[i]
                predicted_return = markov_predictions[i - 1] / 100  # 假设状态对应某一百分比
                y.append(actual_return - predicted_return)
            
            return np.array(X), np.array(y)
        
        except Exception as e:
            logger.error(f"准备LSTM数据时发生错误: {str(e)}")
            raise

    def train_lstm(self):
        """訓練LSTM模型，包含模型保存和載入功能"""
        try:
            logger.info("開始訓練LSTM模型流程")
            
            # 確保模型保存目錄存在
            model_dir = 'models'
            if not os.path.exists(model_dir):
                os.makedirs(model_dir)
                
            # 修改文件命名格式
            model_path = os.path.join(model_dir, f'{self.symbol}_lstm_model.h5')
            weights_path = os.path.join(model_dir, f'{self.symbol}_lstm.weights.h5')
            
            # 準備訓練數據
            X, y = self.prepare_lstm_data()
            train_size = int(len(X) * 0.8)
            X_train, X_val = X[:train_size], X[train_size:]
            y_train, y_val = y[:train_size], y[train_size:]
            
            # 檢查是否存在已訓練的模型
            if os.path.exists(model_path) and os.path.exists(weights_path):
                logger.info("發現已存在的模型，載入模型和權重")
                # 清除之前的自定義對象
                tf.keras.backend.clear_session()
                
                # 重新創建模型
                self.lstm_model = Sequential([
                    LSTM(128, return_sequences=True, input_shape=(X.shape[1], X.shape[2])),
                    Dropout(0.3),
                    LSTM(64, return_sequences=True),
                    Dropout(0.3),
                    LSTM(32),
                    Dropout(0.3),
                    Dense(16, activation='relu'),
                    Dense(1)
                ])
                
                # 使用相同的編譯設置
                self.lstm_model.compile(
                    optimizer=tf.keras.optimizers.Adam(learning_rate=0.001),
                    loss=tf.keras.losses.MeanSquaredError()
                )
                
                # 載入權重
                self.lstm_model.load_weights(weights_path)
                
                # 使用新數據繼續訓練模型
                logger.info("使用新數據更新模型")
                history = self.lstm_model.fit(
                    X_train, y_train,
                    epochs=100,
                    batch_size=32,
                    validation_data=(X_val, y_val),
                    verbose=1,
                    callbacks=[
                        EarlyStopping(
                            monitor='val_loss',
                            patience=10,
                            restore_best_weights=True
                        )
                    ]
                )
            else:
                logger.info("創建新的LSTM模型")
                # 建構新模型
                self.lstm_model = Sequential([
                    LSTM(128, return_sequences=True, input_shape=(X.shape[1], X.shape[2])),
                    Dropout(0.3),
                    LSTM(64, return_sequences=True),
                    Dropout(0.3),
                    LSTM(32),
                    Dropout(0.3),
                    Dense(16, activation='relu'),
                    Dense(1)
                ])
                
                # 編譯模型
                self.lstm_model.compile(
                optimizer=tf.keras.optimizers.Adam(learning_rate=0.001),
                loss=tf.keras.losses.MeanSquaredError(),
                metrics=['mae']
                )
                # 訓練模型
                history = self.lstm_model.fit(
                    X_train, y_train,
                    epochs=100,
                    batch_size=32,
                    validation_data=(X_val, y_val),
                    verbose=1,
                    callbacks=[
                        EarlyStopping(
                            monitor='val_loss',
                            patience=10,
                            restore_best_weights=True
                        )
                    ]
                )

            # 保存訓練歷史
            self.training_history = history.history
            # 計算準確率
            train_loss, train_mae = self.lstm_model.evaluate(X_train, y_train, verbose=0)
            val_loss, val_mae = self.lstm_model.evaluate(X_val, y_val, verbose=0)

            # 保存準確率
            self.accuracy = {
            'train_mae': train_mae,
            'val_mae': val_mae
        }
        
            logger.info(f"訓練集準確率（MAE）：{train_mae:.4f}, 驗證集準確率（MAE）：{val_mae:.4f}")
        
            # 保存更新後的模型和權重
            logger.info("保存模型和權重")
            self.lstm_model.save_weights(weights_path)
            
            # 保存模型訓練信息
            info_path = os.path.join(model_dir, f'{self.symbol}_model_info.json')
            model_info = {
                'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'symbol': self.symbol,
                'input_shape': list(X.shape),
                'training_samples': len(X_train),
                'validation_samples': len(X_val)
            }
            
            with open(info_path, 'w') as f:
                json.dump(model_info, f, indent=4)
                
            logger.info("LSTM模型訓練和保存完成")
            
        except Exception as e:
            logger.error(f"訓練LSTM模型時發生錯誤: {str(e)}")
            raise
    
    #def get_state(change):
    #"""根據變化百分比判斷狀態"""
        #if change > 0.06:
            #return "++"
        #elif change > 0.005:
            #return "+"
        #elif change > -0.005:
            #return "0"
        #elif change > -0.06:
            #return "-"
       # else:
            #return "--"

    @staticmethod
    def sliding_window_markov(data):
    #"""使用滑動視窗生成馬可夫轉移矩陣"""
        WINDOW_SIZE = 20
        def get_state(change):
            if change > 0.06:
                return "++"
            elif change > 0.005:
                return "+"
            elif change > -0.005:
                return "0"
            elif change > -0.06:
                return "-"
            else:
                return "--"

        try:
            if len(data) <= WINDOW_SIZE:
                raise ValueError("數據量不足以計算轉移矩陣")

        # 計算狀態序列
            states = [get_state((data[i + 1] - data[i]) / data[i]) for i in range(len(data) - 1)]

        # 初始化轉移矩陣
            transition_matrix = np.zeros((5, 5))
            state_map = {"++": 0, "+": 1, "0": 2, "-": 3, "--": 4}

        # 滑動視窗計算轉移次數
            for i in range(len(states) - WINDOW_SIZE):
                window_states = states[i:i + WINDOW_SIZE]
            for j in range(WINDOW_SIZE - 1):
                current_state = state_map[window_states[j]]
                next_state = state_map[window_states[j + 1]]
                transition_matrix[current_state, next_state] += 1

        # 將次數轉為機率
            transition_matrix = np.divide(
                transition_matrix,
                transition_matrix.sum(axis=1, keepdims=True),
                where=transition_matrix.sum(axis=1, keepdims=True) != 0
        )

            return transition_matrix

        except Exception as e:
            logger.error(f"計算轉移矩陣時出錯: {e}")
            raise

    def prepare_markov_chain(self):
        """准备马尔可夫链模型"""
        try:
            logger.info("开始准备马尔可夫链模型")
            
            # 计算日收益率
            self.data['Returns'] = self.data['Close'].pct_change()
            
            # 根据收益率判断状态
            def get_state(return_value):
                for state, threshold in self.state_thresholds.items():
                    if return_value > threshold:
                        return state
                return '大跌'
            
            # 使用新的状态判断函数
            self.data['State'] = self.data['Returns'].apply(get_state)
            
            # 创建并初始化转移矩阵（使用numpy数组）
            num_states = len(self.ordered_states)
            self.transition_matrix = np.zeros((num_states, num_states))
            
            # 创建状态到索引的映射
            state_to_index = {state: idx for idx, state in enumerate(self.ordered_states)}
            
            # 计算状态转移次数
            for i in range(len(self.data)-1):
                current_state = self.data['State'].iloc[i]
                next_state = self.data['State'].iloc[i+1]
                current_idx = state_to_index[current_state]
                next_idx = state_to_index[next_state]
                self.transition_matrix[current_idx][next_idx] += 1
            
            # 计算转移概率
            row_sums = self.transition_matrix.sum(axis=1)
            row_sums[row_sums == 0] = 1  # 避免除以零
            self.transition_matrix = self.transition_matrix / row_sums[:, np.newaxis]
            
            logger.info("马尔可夫链模型准备完成")
            return self.transition_matrix
            
        except Exception as e:
            logger.error(f"准备马尔可夫链时发生错误: {str(e)}")
            raise

        #try:
            #logger.info("開始準備馬可夫鏈模型")
            
            # 計算日收益率
            #self.data['Returns'] = self.data['Close'].pct_change()
            
            # 根據收益率判斷狀態
            #def get_state(return_value):
                #for state, threshold in self.state_thresholds.items():
                    #if return_value > threshold:
                        #return state
                #return '大跌'
            
            # 使用新的狀態判斷函數
            #self.data['State'] = self.data['Returns'].apply(get_state)
            
            # 創建並初始化轉移矩陣
            #self.markov_states = pd.DataFrame(
                #0,
                #index=self.ordered_states,
                #columns=self.ordered_states
            #)
            
            # 計算狀態轉移次數
            #for i in range(len(self.data)-1):
                #current_state = self.data['State'].iloc[i]
                #next_state = self.data['State'].iloc[i+1]
                #self.markov_states.loc[current_state, next_state] += 1
            
            # 計算轉移機率
            #self.transition_matrix = self.markov_states.div(
                #self.markov_states.sum(axis=1), axis=0
            #).fillna(0)
            
            # 強制按照 ordered_states 排序
            #self.transition_matrix = self.transition_matrix.reindex(
                #index=self.ordered_states,
                #columns=self.ordered_states
            #)
            
            # 計算狀態統計
            #state_counts = pd.Series(0, index=self.ordered_states)  # 初始化為0
            #value_counts = self.data['State'].value_counts()
            #state_counts.update(value_counts)
            #state_percentages = (state_counts / len(self.data) * 100).round(2)
            
            # 準備狀態統計信息
            ##self.state_stats = OrderedDict()
            ##for state in self.ordered_states:
                ##self.state_stats[state] = {
                    ##'count': int(state_counts[state]),
                    ##'percentage': float(state_percentages[state]),
                    ##'symbol': self.state_symbols[state]
                ##}
                
            #logger.info("馬可夫鏈模型準備完成")
            
        #except Exception as e:
            #logger.error(f"準備馬可夫鏈時發生錯誤: {str(e)}")
            #raise

    def predict_next_day(self):
        """预测下一天的股价"""
        try:
            logger.info("开始进行预测")
            
            # 准备预测数据
            df = self.prepare_features()
            
            # 使用完整的特征列表，确保顺序一致
            features = [
                'Returns', 'Volatility', 
                'Price_MA5_Diff', 'Price_MA20_Diff',
                'RSI', 'BB_width', 
                'MACD_Hist', 'Volume_Ratio',
                'Markov_State'
            ]
            
            # 检查特征是否都存在
            for feature in features:
                if feature not in df.columns:
                    raise ValueError(f"预测时缺少特征: {feature}")
            
            # 取最后 sequence_length 个数据点
            last_sequence = df[features].tail(self.sequence_length)
            
            # 进行特征缩放
            scaled_sequence = self.feature_scaler.transform(last_sequence)
            
            # 重塑数据为 LSTM 所需的形状
            X_pred = scaled_sequence.reshape(1, self.sequence_length, len(features))
            
            # 预测下一天的收益率
            pred_return = self.lstm_model.predict(X_pred)[0][0]
            
            # 将收益率转换为价格
            current_price = float(df['Close'].iloc[-1])
            pred_price = current_price * (1 + pred_return)
            
            # 计算预测区间
            volatility = df['Returns'].tail(20).std()
            confidence_interval = {
                'lower_bound': float(pred_price * (1 - 1.96 * volatility)),
                'upper_bound': float(pred_price * (1 + 1.96 * volatility))
            }
            
            # 获取马尔可夫链预测
            current_state_idx = int(df['Markov_State'].iloc[-1])
            current_state = self.ordered_states[current_state_idx]
            
            # 获取下一状态预测概率
            next_state_probs = self.transition_matrix[current_state_idx]
            next_state_idx = np.argmax(next_state_probs)
            markov_predicted_state = self.ordered_states[next_state_idx]
            
            # 计算状态统计信息
            total_samples = len(df)
            state_counts = np.zeros(len(self.ordered_states))
            for i in range(len(self.ordered_states)):
                state_counts[i] = np.sum(df['Markov_State'] == i)
            
            state_statistics = {}
            for i, state in enumerate(self.ordered_states):
                state_statistics[state] = {
                    'count': int(state_counts[i]),
                    'percentage': float((state_counts[i] / total_samples * 100).round(2)),
                    'symbol': self.state_symbols[state]
                }
            
            # 转换转移矩阵为字典格式
            transition_matrix_dict = {}
            for i, state in enumerate(self.ordered_states):
                transition_matrix_dict[state] = {}
                for j, next_state in enumerate(self.ordered_states):
                    transition_matrix_dict[state][next_state] = float(self.transition_matrix[i, j])
            
            predictions = {
                'prediction_date': (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'),
                'current_price': current_price,
                'predicted_price': float(pred_price),
                'predicted_return': float(pred_return * 100),
                'confidence_interval': confidence_interval,
                'technical_signals': self.get_technical_signals(df),
                'markov_prediction': {
                    'current_state': {
                        'state': current_state,
                        'symbol': self.state_symbols[current_state]
                    },
                    'predicted_state': {
                        'state': markov_predicted_state,
                        'symbol': self.state_symbols[markov_predicted_state]
                    },
                    'transition_matrix': transition_matrix_dict,
                    'state_statistics': state_statistics  # 添加状态统计信息
                }
            }
            
            # 添加训练历史图表
            training_history_img = self.plot_training_history()
            if training_history_img:
                predictions['training_history_image'] = training_history_img
            
            logger.info("预测完成")
            return predictions
            
        except Exception as e:
            logger.error(f"预测过程中发生错误: {str(e)}")
            raise

    def plot_training_history(self):
        """繪製訓練歷史圖表"""
        try:
            plt.figure(figsize=(12, 6))
            
            # 繪製損失曲線
            epochs = range(1, len(self.training_history['loss']) + 1)
            plt.plot(epochs, self.training_history['loss'], 'b--', label='Train', linewidth=2)
            plt.plot(epochs, self.training_history['val_loss'], color='orange', label='Validation', linewidth=2)
            
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.title('LSTM Training History', fontsize=16, pad=20)
            plt.xlabel('Epoch', fontsize=12)
            plt.ylabel('Loss (MSE)', fontsize=12)  # 修改這裡
            plt.legend()
            
            # 保存圖表為base64字符串
            img = io.BytesIO()
            plt.savefig(img, format='png', bbox_inches='tight', dpi=300)
            img.seek(0)
            training_history_img = base64.b64encode(img.getvalue()).decode('utf-8')
            
            plt.close()
            
            return training_history_img
            
        except Exception as e:
            logger.error(f"繪製訓練歷史圖表時發生錯誤: {str(e)}")
            return None

    def get_technical_signals(self, df):
        """獲取技術指標信號"""
        latest = df.iloc[-1]
        
        signals = {
        '均線趨勢指標': '上升趨勢' if latest['MA5'] > latest['MA20'] else '下降趨勢',
        'RSI指標': '買超' if latest['RSI'] > 70 else '超賣' if latest['RSI'] < 30 else '持平',
        'MACD指標': '買進訊號' if latest['MACD'] > latest['Signal_Line'] else '賣出訊號',
        '布林通道': '上軌以上' if latest['Close'] > latest['BB_upper'] else '下軌以下' if latest['Close'] < latest['BB_lower'] else '中間範圍',
        '成交量': '高' if latest['Volume_Ratio'] > 1.5 else '低' if latest['Volume_Ratio'] < 0.5 else '一般'
        }
        
        return signals
    
class StockRecorder:
    def __init__(self, excel_filename='InvestMania準確度紀錄.xlsx'):
        """
        初始化股票記錄器
        :param excel_filename: Excel文件名稱
        """
        self.file_path = excel_filename
        self.excel_filename = excel_filename
        self.headers = [
            '日期', '目前股價', '預測價格', '預期報酬率', '預期區間', 
            '技術指標信號', 'MACD', 'RSI', '均線趨勢', '布林通道', 
            '成交量', '馬可夫鏈分析', '目前狀態', '預測狀態'
        ]
        self._initialize_workbook()

    def _initialize_workbook(self):
        """初始化工作簿和總覽工作表"""
        try:
            if not os.path.exists(self.excel_filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "總覽"
                
                # 設置總覽工作表的標題
                ws['A1'] = '股票代碼'
                ws['B1'] = '最後更新'
                ws['C1'] = '預期報酬'
                ws['D1'] = '預測狀態'
                
                # 設置標題樣式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                wb.save(self.excel_filename)
                
            else:
                wb = load_workbook(self.file_path)
                if "總覽" not in wb.sheetnames:
                    ws = wb.create_sheet("總覽")
                    ws['A1'] = '股票代碼'
                    ws['B1'] = '最後更新'
                    ws['C1'] = '預期報酬'
                    ws['D1'] = '預測狀態'
                    
                    # 設置標題樣式
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    wb.save(self.excel_filename)
                wb.close()
        except Exception as e:
            logger.error(f"初始化Excel文件時發生錯誤: {str(e)}")
            raise

    def ensure_file_exists(self):
        """確保Excel文件存在，如果不存在則創建"""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            # 保留默認工作表並重命名為"總覽"
            ws = wb.active
            ws.title = "總覽"
            self.format_worksheet(ws)  # 格式化總覽工作表
            wb.save(self.file_path)

    def format_worksheet(self, ws):
        """設置工作表的格式"""
        # 設置列寬
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        widths = [12, 10, 12, 12, 15, 15, 8, 8, 10, 10, 8, 12, 10, 10]
        
        for col, width in zip(columns, widths):
            ws.column_dimensions[col].width = width

        # 設置標題行格式
        headers = ['日期', '目前股價', '預測價格', '預期報酬率', '預測區間', '技術指標信號',
                  'MACD', 'RSI', '均線趨勢', '布林通道', '成交量', '馬可夫鏈分析',
                  '目前狀態', '預測狀態']
        
        # 設置表格邊框樣式
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 設置標題行樣式
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        return ws

    def add_prediction_record(self, symbol, prediction_data):
        try:
            wb = load_workbook(self.file_path)
            
            # 確保工作表存在
            if symbol not in wb.sheetnames:
                ws = wb.create_sheet(symbol)
                self.format_worksheet(ws)
            else:
                ws = wb[symbol]

            # 準備新行數據
            row_data = [
                prediction_data.get('prediction_date', ''),
                prediction_data.get('current_price', ''),
                prediction_data.get('predicted_price', ''),
                f"{prediction_data.get('predicted_return', ''):.2f}%",
                f"{prediction_data['confidence_interval']['lower_bound']:.2f}-{prediction_data['confidence_interval']['upper_bound']:.2f}",
                '',  # 技術指標信號 - 保持空白
                prediction_data['technical_signals'].get('MACD指標', ''),
                prediction_data['technical_signals'].get('RSI指標', ''),
                prediction_data['technical_signals'].get('均線趨勢指標', ''),
                prediction_data['technical_signals'].get('布林通道', ''),
                prediction_data['technical_signals'].get('成交量', ''),
                '',  # 馬可夫鏈分析 - 保持空白
                prediction_data['markov_prediction']['current_state']['state'],
                prediction_data['markov_prediction']['predicted_state']['state']
            ]

            # 插入新行
            next_row = ws.max_row + 1
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 移除底色設定，只保留邊框和置中對齊
                cell.font = Font(color="000000")  # 全部使用黑色字體

            # 保存文件
            wb.save(self.file_path)
            return True

        except Exception as e:
            print(f"記錄預測結果時發生錯誤: {str(e)}")
            return False

    def _update_overview(self, workbook, symbol, prediction_data):
        """更新總覽工作表"""
        ws = workbook["總覽"]
        
        # 在總覽中查找或創建該股票的行
        symbol_col = 1
        date_col = 2
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=symbol_col).value == symbol:
                found = True
                update_row = row
                break
        
        if not found:
            update_row = ws.max_row + 1
            ws.cell(row=update_row, column=symbol_col).value = symbol

        # 更新資訊
        ws.cell(row=update_row, column=date_col).value = prediction_data.get('prediction_date', '')
        ws.cell(row=update_row, column=3).value = prediction_data.get('predicted_return', '')
        ws.cell(row=update_row, column=4).value = prediction_data['markov_prediction']['predicted_state']['state']

        # 設置樣式
        for col in range(1, 5):
            cell = ws.cell(row=update_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _format_technical_signals(self, signals):
        """格式化技術指標信號"""
        if not signals:
            return ''
        return '; '.join(f'{k}: {v}' for k, v in signals.items())

@app.route('/')
def welcome():
    return render_template('welcome.html')

@app.route('/predict-form')
def predict_form():
    return render_template('index.html')

@app.route('/instructions')
def instructions():
    return render_template('instructions.html')


@app.route("/predict", methods=["POST"])
def predict():
    symbol = request.form.get("symbol")
    if not symbol:
        return jsonify({"error": "請提供股票代碼"})

    try:
        # 抓取數據
        data = fetch_stock_data(symbol)
        if data is None:
            return jsonify({"error": "無法獲取股票數據"})

        # 創建 StockPredictor 實例
        predictor = StockPredictor(data, symbol)

        # 準備馬可夫鏈
        predictor.prepare_markov_chain()

        # 準備 LSTM 數據並訓練模型
        predictor.train_lstm()

        # 獲取預測結果
        prediction_results = predictor.predict_next_day()

        # 獲取準確率
        accuracy = predictor.accuracy

        # 繪製股價趨勢圖
        trend_image = plot_stock_trend(data)

        # 記錄預測結果到Excel (使用正確的方法名)
        try:
            recorder = StockRecorder()
            recorder.add_prediction_record(symbol, prediction_results)
            logger.info(f"成功記錄 {symbol} 的預測結果")
        except Exception as e:
            logger.error(f"記錄預測結果時發生錯誤: {str(e)}")
            # 記錄錯誤但不中斷流程

        # 返回完整的預測結果
        return jsonify({
            "success": True,
            "predictions": {
                "prediction_date": prediction_results["prediction_date"],
                "current_price": prediction_results["current_price"],
                "predicted_price": prediction_results["predicted_price"],
                "predicted_return": prediction_results["predicted_return"],
                "confidence_interval": prediction_results["confidence_interval"],
                "technical_signals": prediction_results["technical_signals"],
                "markov_prediction": prediction_results["markov_prediction"],
                "training_history_image": prediction_results.get("training_history_image"),
            },
            "trend_image": trend_image,  # 直接將 trend_image 放在外層
            "accuracy": accuracy
        })

    except Exception as e:
        logger.error(f"預測時發生錯誤: {str(e)}")
        return jsonify({"error": str(e)})
    
if __name__ == '__main__':
    app.run(debug=True)