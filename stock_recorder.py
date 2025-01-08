import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import logging
from datetime import datetime
import os

# 設置日誌
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class StockRecorder:
    def __init__(self, excel_filename='InvestMania準確度紀錄.xlsx'):
        """
        初始化股票記錄器
        :param excel_filename: Excel文件名稱
        """
        self.excel_filename = excel_filename
        self.headers = [
            '日期', '目前股價', '預測價格', '預期報酬率', '預期區間', 
            '技術指標信號', 'MACD', 'RSI', '均線趨勢', '布林通道', 
            '成交量', '馬可夫鏈分析', '目前狀態', '預測狀態'
        ]
        
    def _get_or_create_workbook(self):
        """獲取或創建Excel工作簿"""
        try:
            if os.path.exists(self.excel_filename):
                return openpyxl.load_workbook(self.excel_filename)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # 移除默認的工作表
                return wb
        except Exception as e:
            logger.error(f"操作Excel文件時發生錯誤: {str(e)}")
            raise

    def _get_or_create_worksheet(self, wb, symbol):
        """獲取或創建工作表"""
        if symbol in wb.sheetnames:
            ws = wb[symbol]
        else:
            ws = wb.create_sheet(symbol)
            self._setup_worksheet_header(ws)
        return ws

    def _setup_worksheet_header(self, ws):
        """設置工作表表頭"""
        # 添加表頭
        ws.append(self.headers)
        
        # 設置表頭樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 調整列寬
        for i, header in enumerate(self.headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(header) * 1.5, 12)

    def save_prediction(self, predictions, symbol):
        """
        保存預測結果到Excel
        :param predictions: 預測結果字典
        :param symbol: 股票代碼
        """
        try:
            # 獲取或創建工作簿
            wb = self._get_or_create_workbook()
            
            # 獲取或創建工作表
            ws = self._get_or_create_worksheet(wb, symbol)
            
            # 準備數據
            technical_signals = predictions['technical_signals']
            markov_pred = predictions['markov_prediction']
            
            # 構建行數據
            row_data = [
                predictions['prediction_date'],
                predictions['current_price'],
                predictions['predicted_price'],
                f"{predictions['predicted_return']:.2f}%",
                f"{predictions['confidence_interval']['lower_bound']:.2f}-{predictions['confidence_interval']['upper_bound']:.2f}",
                "", # 技術指標信號列保留空白，作為分隔
                technical_signals['MACD指標'],
                technical_signals['RSI指標'],
                technical_signals['均線趨勢指標'],
                technical_signals['布林通道'],
                technical_signals['成交量'],
                "", # 馬可夫鏈分析列保留空白，作為分隔
                f"{markov_pred['current_state']['state']}",
                f"{markov_pred['predicted_state']['state']}"
            ]
            
            # 添加新行
            ws.append(row_data)
            
            # 設置新行的樣式
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center')
            
            # 保存文件
            wb.save(self.excel_filename)
            logger.info(f"已將{symbol}的預測結果保存到 {self.excel_filename}")
            
        except Exception as e:
            logger.error(f"保存預測結果時發生錯誤: {str(e)}")
            raise

    def get_prediction_history(self, symbol):
        """
        獲取特定股票的預測歷史
        :param symbol: 股票代碼
        :return: 預測歷史列表
        """
        try:
            if not os.path.exists(self.excel_filename):
                return []
                
            wb = openpyxl.load_workbook(self.excel_filename, read_only=True)
            if symbol not in wb.sheetnames:
                return []
                
            ws = wb[symbol]
            history = []
            
            # 跳過表頭行
            rows = list(ws.rows)[1:]
            for row in rows:
                row_data = [cell.value for cell in row]
                if len(row_data) >= len(self.headers):
                    history.append(dict(zip(self.headers, row_data)))
            
            return history
            
        except Exception as e:
            logger.error(f"獲取預測歷史時發生錯誤: {str(e)}")
            return []

# 使用示例：
if __name__ == "__main__":
    # 創建記錄器實例
    recorder = StockRecorder()
    
    # 示例預測數據
    sample_predictions = {
        'prediction_date': '2024-12-26',
        'current_price': 1085,
        'predicted_price': 1086.89,
        'predicted_return': 0.17,
        'confidence_interval': {'lower_bound': 1047, 'upper_bound': 1126.37},
        'technical_signals': {
            'MACD指標': '買進訊號',
            'RSI指標': '持平',
            '均線趨勢指標': '上升趨勢',
            '布林通道': '中間範圍',
            '成交量': '低'
        },
        'markov_prediction': {
            'current_state': {'state': '平盤'},
            'predicted_state': {'state': '小跌'}
        }
    }
    
    # 保存預測結果
    recorder.save_prediction(sample_predictions, '2330')
    
    # 獲取預測歷史
    history = recorder.get_prediction_history('2330')
    print(f"預測歷史記錄數: {len(history)}")